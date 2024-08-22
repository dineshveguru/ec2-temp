from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
import pandas as pd
import os
import openpyxl
import time
import re


def extract_total_bed_count(text):
    try:
        pattern = r"(.{0,20})\bbed(?:s|ded)?\b(.{0,20})"
        total_beds = 0
        numbers = []
        for match in re.finditer(pattern, text):
            before_text = match.group(1)
            after_text = match.group(2)
            if "." in before_text:
                before_text = before_text.split(".")[-1]
            before_numbers = re.findall(r"\d+", before_text)
            if before_numbers:
                closest_before_number = int(before_numbers[-1])
                if closest_before_number not in numbers:
                    numbers.append(closest_before_number)
                    total_beds += closest_before_number
                continue
            after_numbers = re.findall(r"\d+", after_text)
            if after_numbers:
                closest_after_number = int(after_numbers[0])
                if closest_after_number not in numbers:
                    numbers.append(closest_after_number)
                    total_beds += closest_after_number
        return total_beds
    except:
        return 0


wb = openpyxl.load_workbook("hospital.xlsx")
ws = wb.active
options = Options()
# options.add_experimental_option("detach", True)
options.add_argument("--headless")
options.add_argument("window-size=1200x600")
ws.cell(row=1, column=18, value="Link")

driver = webdriver.Chrome(options=options)
for i in range(6, 10):
    try:
        name = ws.cell(row=i, column=1).value
        state = ws.cell(row=i, column=2).value
        city = ws.cell(row=i, column=3).value
        driver.get("https://www.google.com/")
        search_box = driver.find_element(By.XPATH, "//textarea[@aria-label='Search']")

        search_box.send_keys(f"{name} {city} {state} beds count")

        search_box.submit()
        try:
            location_element = WebDriverWait(driver, 3).until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "//div[@class='DiqQLb wHYlTd']",
                    )
                )
            )
            location_element.find_element(
                By.XPATH, "//div[@class='mpQYc']//g-raised-button"
            ).click()
        except:
            pass
        link_element = driver.find_element(By.XPATH, "//a[@jsname='UWckNb']")
        link = link_element.get_attribute("href")
        link_element.click()
        elements = driver.find_elements(
            By.XPATH, "//*[contains(translate(text(), 'BED', 'bed'), 'bed')]"
        )
        print("--------hospital name: ", name)

        elements = [i.text.strip() for i in elements]
        elements = list(set(elements))
        whole_string = "\n".join(elements)
        with open("bed_count_data.md", "a") as f:
            f.write("# hospital name: " + name + "\n")
            f.write(whole_string)
            f.write("\n\n")
        bed_count = extract_total_bed_count(whole_string)
        print("bed count: ", bed_count)
        ws.cell(row=i, column=17, value=bed_count)
        ws.cell(row=i, column=18, value=link)
        print(f"{i+1} / {200} completed!")
        time.sleep(0.1)
    except:
        pass

    finally:
        wb.save("hospital.xlsx")
